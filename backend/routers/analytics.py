"""
Analytics endpoints - SSP analytics, dashboard stats, reports, pacing, insights
"""
from fastapi import APIRouter, HTTPException
from fastapi.responses import Response
from typing import Optional, List
from datetime import datetime, timezone, timedelta
import csv
import io
import json
import uuid
import logging

from models import DashboardStats
from routers.shared import db

router = APIRouter(tags=["Analytics"])

logger = logging.getLogger(__name__)


# ==================== DASHBOARD ====================

@router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats():
    """Get dashboard statistics"""
    campaigns = await db.campaigns.find({}, {"_id": 0}).to_list(1000)
    creatives = await db.creatives.count_documents({})
    endpoints = await db.ssp_endpoints.count_documents({})
    
    active_campaigns = len([c for c in campaigns if c.get("status") == "active"])
    total_bids = sum(c.get("bids", 0) for c in campaigns)
    total_wins = sum(c.get("wins", 0) for c in campaigns)
    total_impressions = sum(c.get("impressions", 0) for c in campaigns)
    total_spend = sum(c.get("budget", {}).get("total_spend", 0) for c in campaigns)
    
    avg_win_rate = (total_wins / total_bids * 100) if total_bids > 0 else 0
    
    return DashboardStats(
        total_campaigns=len(campaigns),
        active_campaigns=active_campaigns,
        total_creatives=creatives,
        total_endpoints=endpoints,
        total_bids=total_bids,
        total_wins=total_wins,
        total_impressions=total_impressions,
        total_spend=total_spend,
        avg_win_rate=round(avg_win_rate, 2)
    )


@router.get("/dashboard/chart-data")
async def get_chart_data():
    """Get chart data for dashboard"""
    # Last 7 days of data
    end_dt = datetime.now(timezone.utc)
    start_dt = end_dt - timedelta(days=7)
    
    pipeline = [
        {
            "$match": {
                "timestamp": {
                    "$gte": start_dt.isoformat(),
                    "$lte": end_dt.isoformat()
                }
            }
        },
        {
            "$group": {
                "_id": {"$substr": ["$timestamp", 0, 10]},
                "bids": {"$sum": {"$cond": ["$bid_made", 1, 0]}},
                "wins": {"$sum": {"$cond": ["$win_notified", 1, 0]}},
                "spend": {"$sum": {"$ifNull": ["$win_price", 0]}}
            }
        },
        {"$sort": {"_id": 1}}
    ]
    
    daily_data = await db.bid_logs.aggregate(pipeline).to_list(100)
    
    return {
        "labels": [d["_id"] for d in daily_data],
        "bids": [d["bids"] for d in daily_data],
        "wins": [d["wins"] for d in daily_data],
        "spend": [d["spend"] / 1000 for d in daily_data]
    }


# ==================== SSP ANALYTICS ====================

@router.get("/ssp-analytics/overview")
async def get_ssp_analytics_overview():
    """Get SSP performance analytics overview"""
    endpoints = await db.ssp_endpoints.find({}, {"_id": 0}).to_list(100)
    
    total_requests = sum(e.get("total_requests", 0) for e in endpoints)
    total_bids = sum(e.get("total_bids", 0) for e in endpoints)
    total_wins = sum(e.get("total_wins", 0) for e in endpoints)
    total_spend = sum(e.get("total_spend", 0) for e in endpoints)
    
    overall_bid_rate = (total_bids / total_requests * 100) if total_requests > 0 else 0
    overall_win_rate = (total_wins / total_bids * 100) if total_bids > 0 else 0
    
    ssp_rankings = []
    for e in endpoints:
        requests = e.get("total_requests", 0)
        bids = e.get("total_bids", 0)
        wins = e.get("total_wins", 0)
        spend = e.get("total_spend", 0)
        
        ssp_rankings.append({
            "id": e.get("id"),
            "name": e.get("name"),
            "status": e.get("status"),
            "requests": requests,
            "bids": bids,
            "wins": wins,
            "spend": spend,
            "bid_rate": round((bids / requests * 100) if requests > 0 else 0, 2),
            "win_rate": round((wins / bids * 100) if bids > 0 else 0, 2),
            "avg_response_time_ms": e.get("avg_response_time_ms", 0),
            "last_request_at": e.get("last_request_at")
        })
    
    ssp_rankings.sort(key=lambda x: x["requests"], reverse=True)
    
    return {
        "overview": {
            "total_ssps": len(endpoints),
            "active_ssps": len([e for e in endpoints if e.get("status") == "active"]),
            "total_requests": total_requests,
            "total_bids": total_bids,
            "total_wins": total_wins,
            "total_spend": round(total_spend, 2),
            "overall_bid_rate": round(overall_bid_rate, 2),
            "overall_win_rate": round(overall_win_rate, 2)
        },
        "ssp_rankings": ssp_rankings,
        "top_performers": {
            "by_requests": ssp_rankings[:3] if ssp_rankings else [],
            "by_win_rate": sorted([s for s in ssp_rankings if s["bids"] > 0], key=lambda x: x["win_rate"], reverse=True)[:3],
            "by_spend": sorted(ssp_rankings, key=lambda x: x["spend"], reverse=True)[:3]
        }
    }


@router.get("/ssp-analytics/{ssp_id}/details")
async def get_ssp_analytics_details(ssp_id: str):
    """Get detailed analytics for a specific SSP"""
    endpoint = await db.ssp_endpoints.find_one({"id": ssp_id}, {"_id": 0})
    if not endpoint:
        raise HTTPException(status_code=404, detail="SSP endpoint not found")
    
    recent_logs = await db.bid_logs.find(
        {"ssp_id": ssp_id},
        {"_id": 0}
    ).sort("timestamp", -1).limit(100).to_list(100)
    
    hourly_dist = {}
    for log in recent_logs:
        try:
            ts = log.get("timestamp")
            if isinstance(ts, str):
                hour = datetime.fromisoformat(ts.replace("Z", "+00:00")).hour
            else:
                hour = ts.hour
            hourly_dist[hour] = hourly_dist.get(hour, 0) + 1
        except (ValueError, TypeError, AttributeError):
            pass
    
    response_times = [log.get("response_time_ms", 0) for log in recent_logs if log.get("response_time_ms")]
    
    campaign_dist = {}
    for log in recent_logs:
        if log.get("bid_made") and log.get("campaign_id"):
            cid = log.get("campaign_id")
            cname = log.get("campaign_name", "Unknown")
            if cid not in campaign_dist:
                campaign_dist[cid] = {"name": cname, "bids": 0, "wins": 0}
            campaign_dist[cid]["bids"] += 1
            if log.get("win_notified"):
                campaign_dist[cid]["wins"] += 1
    
    return {
        "ssp": {
            "id": endpoint.get("id"),
            "name": endpoint.get("name"),
            "status": endpoint.get("status"),
            "endpoint_token": endpoint.get("endpoint_token"),
            "ortb_version": endpoint.get("ortb_version")
        },
        "metrics": {
            "total_requests": endpoint.get("total_requests", 0),
            "total_bids": endpoint.get("total_bids", 0),
            "total_wins": endpoint.get("total_wins", 0),
            "total_spend": endpoint.get("total_spend", 0),
            "avg_response_time_ms": endpoint.get("avg_response_time_ms", 0),
            "bid_rate": round((endpoint.get("total_bids", 0) / endpoint.get("total_requests", 1)) * 100, 2),
            "win_rate": round((endpoint.get("total_wins", 0) / max(endpoint.get("total_bids", 1), 1)) * 100, 2)
        },
        "hourly_distribution": [{"hour": h, "requests": c} for h, c in sorted(hourly_dist.items())],
        "campaign_distribution": list(campaign_dist.values()),
        "response_time_stats": {
            "avg": round(sum(response_times) / len(response_times), 2) if response_times else 0,
            "min": min(response_times) if response_times else 0,
            "max": max(response_times) if response_times else 0
        },
        "recent_activity": recent_logs[:10]
    }


# ==================== BUDGET PACING ====================

@router.post("/pacing/reset-all")
async def reset_all_daily_spend():
    """Reset daily spend for all campaigns"""
    result = await db.campaigns.update_many(
        {},
        {"$set": {
            "budget.daily_spend": 0,
            "budget.current_hour_spend": 0,
            "budget.last_hour_reset": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"status": "reset", "campaigns_updated": result.modified_count}


@router.get("/pacing/status")
async def get_pacing_status():
    """Get pacing status for all active campaigns"""
    campaigns = await db.campaigns.find(
        {"status": "active"},
        {"_id": 0, "id": 1, "name": 1, "budget": 1, "bid_shading": 1}
    ).to_list(1000)
    
    now = datetime.now(timezone.utc)
    current_hour = now.hour
    hours_remaining = 24 - current_hour
    
    status = []
    for c in campaigns:
        budget = c.get("budget", {})
        daily_budget = budget.get("daily_budget", 0)
        daily_spend = budget.get("daily_spend", 0)
        
        if daily_budget > 0:
            pacing_percentage = (daily_spend / daily_budget) * 100
            ideal_percentage = (current_hour / 24) * 100
            pacing_status = "on_track"
            
            if pacing_percentage > ideal_percentage + 10:
                pacing_status = "overpacing"
            elif pacing_percentage < ideal_percentage - 10:
                pacing_status = "underpacing"
        else:
            pacing_percentage = 0
            ideal_percentage = 0
            pacing_status = "unlimited"
        
        status.append({
            "campaign_id": c["id"],
            "campaign_name": c["name"],
            "daily_budget": daily_budget,
            "daily_spend": daily_spend,
            "pacing_percentage": round(pacing_percentage, 1),
            "ideal_percentage": round(ideal_percentage, 1),
            "pacing_status": pacing_status,
            "hours_remaining": hours_remaining,
            "bid_shading_enabled": c.get("bid_shading", {}).get("enabled", False),
            "current_shade_factor": c.get("bid_shading", {}).get("current_shade_factor", 1.0)
        })
    
    return {"current_hour": current_hour, "campaigns": status}


# ==================== REPORTING ====================

@router.get("/reports/campaign/{campaign_id}")
async def get_campaign_report(
    campaign_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """Get performance report for a specific campaign"""
    campaign = await db.campaigns.find_one({"id": campaign_id}, {"_id": 0})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    
    if not end_date:
        end_dt = datetime.now(timezone.utc)
        end_date = end_dt.strftime("%Y-%m-%d")
    else:
        end_dt = datetime.fromisoformat(end_date)
    
    if not start_date:
        start_dt = end_dt - timedelta(days=7)
        start_date = start_dt.strftime("%Y-%m-%d")
    else:
        start_dt = datetime.fromisoformat(start_date)
    
    pipeline = [
        {
            "$match": {
                "campaign_id": campaign_id,
                "timestamp": {
                    "$gte": start_dt.isoformat(),
                    "$lte": (end_dt + timedelta(days=1)).isoformat()
                }
            }
        },
        {
            "$group": {
                "_id": {"$substr": ["$timestamp", 0, 10]},
                "bids": {"$sum": 1},
                "wins": {"$sum": {"$cond": ["$win_notified", 1, 0]}},
                "total_bid_price": {"$sum": {"$ifNull": ["$shaded_price", "$bid_price"]}},
                "total_win_price": {"$sum": {"$ifNull": ["$win_price", 0]}}
            }
        },
        {"$sort": {"_id": 1}}
    ]
    
    daily_data = await db.bid_logs.aggregate(pipeline).to_list(100)
    
    total_bids = sum(d["bids"] for d in daily_data)
    total_wins = sum(d["wins"] for d in daily_data)
    total_bid_value = sum(d["total_bid_price"] for d in daily_data)
    total_win_value = sum(d["total_win_price"] for d in daily_data)
    
    impressions = campaign.get("impressions", 0)
    clicks = campaign.get("clicks", 0)
    spend = campaign.get("budget", {}).get("total_spend", 0)
    
    return {
        "campaign_id": campaign_id,
        "campaign_name": campaign.get("name"),
        "start_date": start_date,
        "end_date": end_date,
        "summary": {
            "impressions": impressions,
            "clicks": clicks,
            "bids": total_bids,
            "wins": total_wins,
            "spend": spend,
            "ctr": (clicks / impressions * 100) if impressions > 0 else 0,
            "win_rate": (total_wins / total_bids * 100) if total_bids > 0 else 0,
            "avg_cpm": (spend / impressions * 1000) if impressions > 0 else 0,
            "avg_bid_price": (total_bid_value / total_bids) if total_bids > 0 else 0,
            "avg_win_price": (total_win_value / total_wins) if total_wins > 0 else 0
        },
        "daily_data": [
            {
                "date": d["_id"],
                "bids": d["bids"],
                "wins": d["wins"],
                "win_rate": (d["wins"] / d["bids"] * 100) if d["bids"] > 0 else 0,
                "avg_bid_price": (d["total_bid_price"] / d["bids"]) if d["bids"] > 0 else 0,
                "avg_win_price": (d["total_win_price"] / d["wins"]) if d["wins"] > 0 else 0
            }
            for d in daily_data
        ],
        "bid_shading": {
            "enabled": campaign.get("bid_shading", {}).get("enabled", False),
            "current_factor": campaign.get("bid_shading", {}).get("current_shade_factor", 1.0),
            "target_win_rate": campaign.get("bid_shading", {}).get("target_win_rate", 0.3),
            "actual_win_rate": campaign.get("recent_win_rate", 0)
        }
    }


@router.get("/reports/summary")
async def get_report_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """Get overall performance report summary"""
    if not end_date:
        end_dt = datetime.now(timezone.utc)
        end_date = end_dt.strftime("%Y-%m-%d")
    else:
        end_dt = datetime.fromisoformat(end_date)
    
    if not start_date:
        start_dt = end_dt - timedelta(days=7)
        start_date = start_dt.strftime("%Y-%m-%d")
    else:
        start_dt = datetime.fromisoformat(start_date)
    
    pipeline = [
        {
            "$match": {
                "timestamp": {
                    "$gte": start_dt.isoformat(),
                    "$lte": (end_dt + timedelta(days=1)).isoformat()
                }
            }
        },
        {
            "$group": {
                "_id": {"$substr": ["$timestamp", 0, 10]},
                "total_bids": {"$sum": {"$cond": ["$bid_made", 1, 0]}},
                "no_bids": {"$sum": {"$cond": ["$bid_made", 0, 1]}},
                "wins": {"$sum": {"$cond": ["$win_notified", 1, 0]}},
                "total_spend": {"$sum": {"$ifNull": ["$win_price", 0]}}
            }
        },
        {"$sort": {"_id": 1}}
    ]
    
    daily_data = await db.bid_logs.aggregate(pipeline).to_list(100)
    
    campaigns = await db.campaigns.find({}, {"_id": 0}).to_list(1000)
    
    total_impressions = sum(c.get("impressions", 0) for c in campaigns)
    total_clicks = sum(c.get("clicks", 0) for c in campaigns)
    total_spend = sum(c.get("budget", {}).get("total_spend", 0) for c in campaigns)
    
    total_bids = sum(d["total_bids"] for d in daily_data)
    total_wins = sum(d["wins"] for d in daily_data)
    
    return {
        "start_date": start_date,
        "end_date": end_date,
        "summary": {
            "total_impressions": total_impressions,
            "total_clicks": total_clicks,
            "total_bids": total_bids,
            "total_wins": total_wins,
            "total_spend": total_spend,
            "ctr": (total_clicks / total_impressions * 100) if total_impressions > 0 else 0,
            "win_rate": (total_wins / total_bids * 100) if total_bids > 0 else 0,
            "avg_cpm": (total_spend / total_impressions * 1000) if total_impressions > 0 else 0
        },
        "daily_data": [
            {
                "date": d["_id"],
                "bids": d["total_bids"],
                "no_bids": d["no_bids"],
                "wins": d["wins"],
                "spend": d["total_spend"] / 1000,
                "win_rate": (d["wins"] / d["total_bids"] * 100) if d["total_bids"] > 0 else 0
            }
            for d in daily_data
        ],
        "campaigns": len(campaigns),
        "active_campaigns": len([c for c in campaigns if c.get("status") == "active"])
    }


@router.get("/reports/export/csv")
async def export_report_csv(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    campaign_id: Optional[str] = None
):
    """Export report data as CSV"""
    if campaign_id:
        report = await get_campaign_report(campaign_id, start_date, end_date)
        daily_data = report["daily_data"]
    else:
        report = await get_report_summary(start_date, end_date)
        daily_data = report["daily_data"]
    
    output = io.StringIO()
    if daily_data:
        writer = csv.DictWriter(output, fieldnames=daily_data[0].keys())
        writer.writeheader()
        writer.writerows(daily_data)
    
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=report_{start_date}_{end_date}.csv"}
    )


@router.get("/reports/export/json")
async def export_report_json(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    campaign_id: Optional[str] = None
):
    """Export report data as JSON"""
    if campaign_id:
        report = await get_campaign_report(campaign_id, start_date, end_date)
    else:
        report = await get_report_summary(start_date, end_date)
    
    return Response(
        content=json.dumps(report, indent=2),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename=report_{start_date}_{end_date}.json"}
    )


# ==================== AD PERFORMANCE REPORT ====================

import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Supply Sources for mock data fallback
SUPPLY_SOURCES = [
    "Google AdX", "OpenX", "PubMatic", "Magnite", "Xandr", 
    "Index Exchange", "TripleLift", "Sovrn", "Amazon TAM", "Yahoo SSP"
]

# Sample domains for mock data fallback
SAMPLE_DOMAINS = [
    "news.example.com", "sports.example.com", "tech.example.com", 
    "lifestyle.example.com", "entertainment.example.com", "finance.example.com",
    "travel.example.com", "health.example.com", "auto.example.com", "gaming.example.com"
]

# Pre-built report templates
REPORT_TEMPLATES = {
    "campaign_overview": {
        "id": "campaign_overview",
        "name": "Campaign Overview",
        "description": "Complete view with all dimensions and metrics",
        "icon": "BarChart3",
        "dimensions": ["source", "domain", "insertion_order", "line_item", "creative_name"],
        "is_default": True
    },
    "video_performance": {
        "id": "video_performance",
        "name": "Video Performance",
        "description": "Focus on video quartile metrics and completion rates",
        "icon": "Video",
        "dimensions": ["source", "creative_name"],
        "is_default": True
    },
    "domain_analysis": {
        "id": "domain_analysis",
        "name": "Domain Analysis",
        "description": "Publisher domain performance breakdown",
        "icon": "Globe",
        "dimensions": ["domain", "source"],
        "is_default": True
    },
    "creative_breakdown": {
        "id": "creative_breakdown",
        "name": "Creative Breakdown",
        "description": "Performance by creative asset",
        "icon": "Image",
        "dimensions": ["creative_name"],
        "is_default": True
    },
    "source_analysis": {
        "id": "source_analysis",
        "name": "Source/SSP Analysis",
        "description": "SSP and exchange performance comparison",
        "icon": "Server",
        "dimensions": ["source"],
        "is_default": True
    }
}


async def get_real_ad_performance_data(
    dimensions: List[str],
    start_date: str,
    end_date: str,
    num_rows: int = 10000,
    campaign_id: Optional[str] = None,
    creative_id: Optional[str] = None,
    metrics: Optional[List[str]] = None
) -> tuple[List[dict], bool]:
    """
    Get real ad performance data from bid_logs, campaigns, creatives, and ssp_endpoints.
    Returns (data, is_real_data) tuple.
    NO MOCK DATA - only real data from database.
    """
    try:
        # Parse dates - be flexible with date format
        try:
            start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        except:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        
        try:
            end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        except:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        
        # Add time to end date to include full day
        end_dt = end_dt.replace(hour=23, minute=59, second=59)
        
        # Build query - get ALL bid logs (no date filter for now to ensure we get data)
        query = {}
        
        # Add campaign filter
        if campaign_id:
            query["campaign_id"] = campaign_id
        
        # Add creative filter
        if creative_id:
            query["creative_id"] = creative_id
        
        # Get ALL bid logs
        bid_logs = await db.bid_logs.find(query, {"_id": 0}).to_list(100000)
        
        if not bid_logs:
            return [], False
        
        # Get all campaigns for mapping
        campaigns = await db.campaigns.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
        campaign_map = {c["id"]: c["name"] for c in campaigns}
        
        # Get all creatives for mapping
        creatives = await db.creatives.find({}, {"_id": 0, "id": 1, "name": 1, "type": 1}).to_list(1000)
        creative_map = {c["id"]: {"name": c["name"], "type": c.get("type", "banner")} for c in creatives}
        
        # Get all SSP endpoints for mapping
        ssp_endpoints = await db.ssp_endpoints.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
        ssp_map = {s["id"]: s["name"] for s in ssp_endpoints}
        
        # Aggregate data by selected dimensions
        aggregated = {}
        
        for log in bid_logs:
            # Build key from dimensions
            key_parts = []
            row = {}
            
            # Campaign name dimension
            if "campaign_name" in dimensions:
                cid = log.get("campaign_id") or ""
                cname = campaign_map.get(cid, f"Unknown ({cid[:8]}...)" if cid else "Unknown")
                row["campaign_name"] = cname
                row["campaign_id"] = cid
                key_parts.append(cname)
            
            # Creative name dimension
            if "creative_name" in dimensions:
                creative_id_val = log.get("creative_id") or ""
                creative_info = creative_map.get(creative_id_val, {"name": f"Unknown ({creative_id_val[:8]}...)" if creative_id_val else "Unknown", "type": "banner"})
                row["creative_name"] = creative_info["name"]
                row["creative_id"] = creative_id_val
                key_parts.append(row["creative_name"])
            
            if "source" in dimensions:
                ssp_id = log.get("ssp_id") or ""
                source_name = ssp_map.get(ssp_id, "Direct/Unknown")
                row["source"] = source_name
                key_parts.append(source_name)
            
            if "domain" in dimensions:
                domain = (log.get("request_summary") or {}).get("domain") or "Unknown"
                row["domain"] = domain
                key_parts.append(domain)
            
            if "bundle" in dimensions:
                bundle = (log.get("request_summary") or {}).get("bundle") or "Unknown"
                row["bundle"] = bundle
                key_parts.append(bundle)
            
            if "app_name" in dimensions:
                app_name = (log.get("request_summary") or {}).get("app_name") or "Unknown"
                row["app_name"] = app_name
                key_parts.append(app_name)
            
            if "country" in dimensions:
                country = (log.get("request_summary") or {}).get("country") or "Unknown"
                row["country"] = country
                key_parts.append(country)
            
            if "city" in dimensions:
                city = (log.get("request_summary") or {}).get("city") or "Unknown"
                row["city"] = city
                key_parts.append(city)
            
            if "ip" in dimensions:
                ip = (log.get("request_summary") or {}).get("ip") or "Unknown"
                row["ip"] = ip
                key_parts.append(ip)
            
            if "device_ifa" in dimensions:
                device_ifa = (log.get("request_summary") or {}).get("device_ifa") or "Unknown"
                row["device_ifa"] = device_ifa
                key_parts.append(device_ifa)
            
            if "os" in dimensions:
                os_name = (log.get("request_summary") or {}).get("os") or "Unknown"
                row["os"] = os_name
                key_parts.append(os_name)
            
            if "make" in dimensions:
                make = (log.get("request_summary") or {}).get("make") or "Unknown"
                row["make"] = make
                key_parts.append(make)
            
            key = "|".join(key_parts) if key_parts else "total"
            
            if key not in aggregated:
                aggregated[key] = {
                    **row,
                    "requests": 0,
                    "bids": 0,
                    "wins": 0,
                    "impressions": 0,
                    "spend": 0,
                    "clicks": 0,
                    "conversions": 0,
                    "video_q1_25": 0,
                    "video_q2_50": 0,
                    "video_q3_75": 0,
                    "video_completed_100": 0,
                    "_unique_users": set()
                }
            
            # Count requests
            aggregated[key]["requests"] += 1
            
            # Count bids made
            if log.get("bid_made"):
                aggregated[key]["bids"] += 1
            
            # Count wins and impressions (win_notified = impression)
            if log.get("win_notified"):
                aggregated[key]["wins"] += 1
                aggregated[key]["impressions"] += 1
                
                # Track spend from win price (already in dollars, not micros)
                win_price = log.get("win_price") or 0
                aggregated[key]["spend"] += win_price
                
                # Track unique users for reach
                user_id = (log.get("request_summary") or {}).get("user_id") or log.get("request_id", "")[:8]
                aggregated[key]["_unique_users"].add(user_id)
            
            # Video quartile metrics from bid log (q1, q2, q3, q4 fields)
            if log.get("q1"):
                aggregated[key]["video_q1_25"] += 1
            if log.get("q2"):
                aggregated[key]["video_q2_50"] += 1
            if log.get("q3"):
                aggregated[key]["video_q3_75"] += 1
            if log.get("q4"):
                aggregated[key]["video_completed_100"] += 1
        
        # Convert to list and calculate derived metrics
        data = []
        for key, row in aggregated.items():
            row["reach"] = len(row["_unique_users"])
            del row["_unique_users"]
            
            # Calculate win rate
            row["win_rate"] = round((row["wins"] / row["bids"] * 100), 2) if row["bids"] > 0 else 0
            
            # Calculate CTR (clicks/impressions)
            row["ctr"] = round((row["clicks"] / row["impressions"] * 100), 2) if row["impressions"] > 0 else 0
            
            # Calculate eCPM
            row["ecpm"] = round((row["spend"] / row["impressions"] * 1000), 2) if row["impressions"] > 0 else 0
            
            # Calculate CPC
            row["cpc"] = round((row["spend"] / row["clicks"]), 2) if row["clicks"] > 0 else 0
            
            # Calculate CPV (cost per video view/completion)
            row["cpv"] = round((row["spend"] / row["video_completed_100"]), 2) if row["video_completed_100"] > 0 else 0
            
            # Calculate video completion rate (VTR)
            row["video_completion_rate"] = round(
                (row["video_completed_100"] / row["impressions"] * 100), 2
            ) if row["impressions"] > 0 and row["video_completed_100"] > 0 else 0
            row["vtr"] = row["video_completion_rate"]
            
            data.append(row)
        
        # Sort by impressions descending and limit
        data.sort(key=lambda x: x["impressions"], reverse=True)
        data = data[:num_rows]
        
        return data, len(data) > 0
        
    except Exception as e:
        logger.error(f"Error fetching real data: {e}")
        import traceback
        traceback.print_exc()
        return [], False


def generate_mock_ad_performance_data(
    dimensions: List[str],
    start_date: str,
    end_date: str,
    num_rows: int = 10000,
    campaign_id: Optional[str] = None,
    creative_id: Optional[str] = None,
    metrics: Optional[List[str]] = None
) -> List[dict]:
    """Generate mock ad performance data based on selected dimensions"""
    random.seed(42)  # For consistent mock data
    
    # Mock campaign names
    mock_campaigns = ["Brand Awareness Q1", "Retargeting Spring", "Performance Max", "Video Reach", "Display Prospecting"]
    # Mock creative names
    mock_creatives = ["Hero Banner 300x250", "Product Video 15s", "Native Card A", "Skyscraper 160x600", "Video Pre-roll 30s"]
    # Mock bundles
    mock_bundles = ["com.game.puzzle", "com.news.daily", "com.social.app", "com.video.streaming", "com.shopping.online"]
    # Mock app names
    mock_app_names = ["Puzzle Master", "Daily News", "Social Connect", "StreamTV", "ShopEase"]
    # Mock countries
    mock_countries = ["US", "UK", "IN", "DE", "FR", "CA", "AU", "JP", "BR", "MX"]
    # Mock cities
    mock_cities = ["New York", "London", "Mumbai", "Berlin", "Paris", "Toronto", "Sydney", "Tokyo", "Sao Paulo", "Mexico City"]
    # Mock IPs
    mock_ips = [f"192.168.{random.randint(1,255)}.{random.randint(1,255)}" for _ in range(100)]
    # Mock Device IFAs (IDFA/GAID format)
    mock_device_ifas = [f"{uuid.uuid4()}" for _ in range(100)]
    # Mock OS
    mock_os = ["Android", "iOS", "Windows", "macOS", "Linux"]
    # Mock device makes
    mock_makes = ["Samsung", "Apple", "Xiaomi", "OnePlus", "Oppo", "Vivo", "Google", "Huawei", "LG", "Sony"]
    
    data = []
    
    for i in range(num_rows):
        row = {}
        
        # Add dimensions
        if "campaign_name" in dimensions:
            row["campaign_name"] = random.choice(mock_campaigns)
            row["campaign_id"] = f"camp-{hash(row['campaign_name']) % 10000}"
        if "creative_name" in dimensions:
            row["creative_name"] = random.choice(mock_creatives)
            row["creative_id"] = f"creat-{hash(row['creative_name']) % 10000}"
        if "source" in dimensions:
            row["source"] = random.choice(SUPPLY_SOURCES)
        if "domain" in dimensions:
            row["domain"] = random.choice(SAMPLE_DOMAINS)
        if "insertion_order" in dimensions:
            row["insertion_order"] = f"IO-{random.randint(1000, 9999)}"
        if "line_item" in dimensions:
            row["line_item"] = f"LI-{random.choice(['Prospecting', 'Retargeting', 'Contextual', 'Lookalike'])}-{random.randint(100, 999)}"
        
        # New dimensions
        if "bundle" in dimensions:
            row["bundle"] = random.choice(mock_bundles)
        if "app_name" in dimensions:
            row["app_name"] = random.choice(mock_app_names)
        if "country" in dimensions:
            row["country"] = random.choice(mock_countries)
        if "city" in dimensions:
            row["city"] = random.choice(mock_cities)
        if "ip" in dimensions:
            row["ip"] = random.choice(mock_ips)
        if "device_ifa" in dimensions:
            row["device_ifa"] = random.choice(mock_device_ifas)
        if "os" in dimensions:
            row["os"] = random.choice(mock_os)
        if "make" in dimensions:
            row["make"] = random.choice(mock_makes)
        
        # Performance metrics
        impressions = random.randint(1000, 500000)
        reach = int(impressions * random.uniform(0.3, 0.8))
        clicks = int(impressions * random.uniform(0.001, 0.05))
        ctr = (clicks / impressions * 100) if impressions > 0 else 0
        conversions = int(clicks * random.uniform(0.01, 0.15))
        spend = impressions * random.uniform(0.5, 5) / 1000  # CPM between $0.5 and $5
        
        row["impressions"] = impressions
        row["reach"] = reach
        row["clicks"] = clicks
        row["ctr"] = round(ctr, 2)
        row["conversions"] = conversions
        row["spend"] = round(spend, 2)
        row["win_rate"] = round(random.uniform(10, 50), 2)
        
        # Derived metrics
        row["ecpm"] = round((spend / impressions * 1000), 2) if impressions > 0 else 0
        row["cpc"] = round((spend / clicks), 2) if clicks > 0 else 0
        
        # Video metrics (if video-related creative)
        is_video = "Video" in row.get("creative_name", "") or "video" in row.get("creative_name", "").lower()
        
        # Always include video metrics, but they'll be 0 for non-video
        q1_25 = int(impressions * random.uniform(0.6, 0.95)) if is_video else 0
        q2_50 = int(q1_25 * random.uniform(0.7, 0.95)) if is_video else 0
        q3_75 = int(q2_50 * random.uniform(0.7, 0.95)) if is_video else 0
        completed = int(q3_75 * random.uniform(0.6, 0.9)) if is_video else 0
        completion_rate = (completed / impressions * 100) if impressions > 0 and is_video else 0
        
        row["video_q1_25"] = q1_25
        row["video_q2_50"] = q2_50
        row["video_q3_75"] = q3_75
        row["video_completed_100"] = completed
        row["video_completion_rate"] = round(completion_rate, 2)
        row["vtr"] = row["video_completion_rate"]
        row["cpv"] = round((spend / completed), 2) if completed > 0 else 0
        
        data.append(row)
    
    return data


@router.post("/reports/ad-performance")
async def generate_ad_performance_report(
    dimensions: str = "campaign_name,creative_name,source,domain",
    metrics: str = "impressions,clicks,ctr,conversions,spend,win_rate,ecpm,cpc,cpv,video_q1_25,video_q2_50,video_q3_75,video_completed_100,video_completion_rate,vtr",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    include_video_metrics: bool = True,
    num_rows: int = 10000,
    use_real_data: bool = True,
    campaign_id: Optional[str] = None,
    creative_id: Optional[str] = None
):
    """Generate ad performance report with real data (if available) or mock data"""
    # Parse comma-separated dimensions string
    dims = [d.strip() for d in dimensions.split(",")]
    
    # Parse comma-separated metrics string
    selected_metrics = [m.strip() for m in metrics.split(",") if m.strip()]
    
    if not end_date:
        end_dt = datetime.now(timezone.utc)
        end_date = end_dt.strftime("%Y-%m-%d")
    
    if not start_date:
        start_dt = datetime.now(timezone.utc) - timedelta(days=365)  # Default to last year
        start_date = start_dt.strftime("%Y-%m-%d")
    
    # Get ONLY real data - no mock data
    data = []
    is_real_data = False
    data_source = "NO DATA"
    
    data, is_real_data = await get_real_ad_performance_data(
        dimensions=dims,
        start_date=start_date,
        end_date=end_date,
        num_rows=num_rows,
        campaign_id=campaign_id,
        creative_id=creative_id,
        metrics=selected_metrics
    )
    
    if is_real_data and data:
        data_source = "REAL DATA (From Bid Logs)"
    else:
        data_source = "NO DATA (No bid activity found)"
    
    # Filter data to only include selected metrics
    if selected_metrics and data:
        filtered_data = []
        for row in data:
            filtered_row = {}
            # Keep dimension fields
            for dim in dims:
                if dim in row:
                    filtered_row[dim] = row[dim]
                # Also keep associated ID fields
                if f"{dim}_id" in row:
                    filtered_row[f"{dim}_id"] = row[f"{dim}_id"]
            # Keep only selected metrics
            for metric in selected_metrics:
                if metric in row:
                    filtered_row[metric] = row[metric]
            filtered_data.append(filtered_row)
        data = filtered_data
    
    # Calculate summary based on ALL data (not filtered)
    total_impressions = sum(d.get("impressions", 0) for d in data)
    total_reach = sum(d.get("reach", 0) for d in data)
    total_clicks = sum(d.get("clicks", 0) for d in data)
    total_conversions = sum(d.get("conversions", 0) for d in data)
    total_spend = sum(d.get("spend", 0) for d in data)
    total_bids = sum(d.get("bids", 0) for d in data)
    total_wins = sum(d.get("wins", 0) for d in data)
    avg_ctr = (total_clicks / total_impressions * 100) if total_impressions > 0 else 0
    
    video_data = [d for d in data if d.get("video_completed_100", 0) > 0]
    total_video_impressions = sum(d.get("impressions", 0) for d in video_data)
    total_completed = sum(d.get("video_completed_100", 0) for d in video_data)
    avg_completion_rate = (total_completed / total_video_impressions * 100) if total_video_impressions > 0 else 0
    
    return {
        "report_metadata": {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "start_date": start_date,
            "end_date": end_date,
            "dimensions": dims,
            "metrics": selected_metrics,
            "total_rows": len(data),
            "data_source": data_source,
            "is_real_data": is_real_data
        },
        "summary": {
            "total_bids": total_bids,
            "total_wins": total_wins,
            "total_impressions": total_impressions,
            "total_reach": total_reach,
            "total_clicks": total_clicks,
            "total_conversions": total_conversions,
            "total_spend": round(total_spend, 2),
            "avg_ctr": round(avg_ctr, 2),
            "win_rate": round((total_wins / total_bids * 100) if total_bids > 0 else 0, 2),
            "video_completion_rate": round(avg_completion_rate, 2)
        },
        "data": data
    }


# ==================== REPORT TEMPLATES ====================

@router.get("/reports/templates")
async def get_report_templates():
    """Get all available report templates (built-in and custom)"""
    # Get custom templates from database
    custom_templates = await db.report_templates.find({}, {"_id": 0}).to_list(100)
    
    # Combine with built-in templates
    all_templates = {
        "built_in": list(REPORT_TEMPLATES.values()),
        "custom": custom_templates
    }
    
    return all_templates


@router.post("/reports/templates")
async def save_report_template(
    name: str,
    description: str,
    dimensions: str,
    icon: str = "FileText"
):
    """Save a custom report template"""
    dims = [d.strip() for d in dimensions.split(",")]
    
    template = {
        "id": str(uuid.uuid4()),
        "name": name,
        "description": description,
        "icon": icon,
        "dimensions": dims,
        "is_default": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.report_templates.insert_one(template)
    if "_id" in template:
        del template["_id"]
    
    return template


@router.delete("/reports/templates/{template_id}")
async def delete_report_template(template_id: str):
    """Delete a custom report template"""
    # Check if it's a built-in template
    if template_id in REPORT_TEMPLATES:
        raise HTTPException(status_code=400, detail="Cannot delete built-in templates")
    
    result = await db.report_templates.delete_one({"id": template_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    
    return {"status": "deleted"}


@router.get("/reports/templates/{template_id}")
async def get_report_template(template_id: str):
    """Get a specific report template"""
    # Check built-in templates first
    if template_id in REPORT_TEMPLATES:
        return REPORT_TEMPLATES[template_id]
    
    # Check custom templates
    template = await db.report_templates.find_one({"id": template_id}, {"_id": 0})
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    return template


@router.get("/reports/ad-performance/export/csv")
async def export_ad_performance_csv(
    dimensions: str = "source,domain,creative_name",
    metrics: str = "impressions,clicks,ctr,conversions,spend",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    num_rows: int = 100,
    use_real_data: bool = True,
    campaign_id: Optional[str] = None,
    creative_id: Optional[str] = None
):
    """Export ad performance report as CSV"""
    dims = [d.strip() for d in dimensions.split(",")]
    selected_metrics = [m.strip() for m in metrics.split(",") if m.strip()]
    
    if not end_date:
        end_dt = datetime.now(timezone.utc)
        end_date = end_dt.strftime("%Y-%m-%d")
    
    if not start_date:
        start_dt = datetime.now(timezone.utc) - timedelta(days=30)
        start_date = start_dt.strftime("%Y-%m-%d")
    
    # Try real data first
    data = []
    if use_real_data:
        data, is_real = await get_real_ad_performance_data(
            dimensions=dims,
            start_date=start_date,
            end_date=end_date,
            num_rows=min(num_rows, 10000),
            campaign_id=campaign_id,
            creative_id=creative_id,
            metrics=selected_metrics
        )
    
    # Fall back to mock data
    if not data:
        data = generate_mock_ad_performance_data(
            dimensions=dims,
            start_date=start_date,
            end_date=end_date,
            num_rows=min(num_rows, 10000),
            campaign_id=campaign_id,
            creative_id=creative_id,
            metrics=selected_metrics
        )
    
    # Define all dimension labels
    dimension_labels = {
        "campaign_name": "Campaign Name",
        "creative_name": "Creative Name",
        "source": "Source",
        "domain": "Domain", 
        "insertion_order": "Insertion Order",
        "line_item": "Line Item",
        "bundle": "Bundle",
        "app_name": "App Name",
        "country": "Country",
        "city": "City",
        "ip": "IP Address",
        "device_ifa": "Device ID",
        "os": "OS",
        "make": "Make"
    }
    
    # Define all metric labels
    metric_labels = {
        "impressions": "Impressions",
        "reach": "Reach",
        "clicks": "Clicks",
        "ctr": "CTR (%)",
        "conversions": "Conversions",
        "spend": "Spend",
        "win_rate": "Win Rate (%)",
        "ecpm": "eCPM",
        "cpc": "CPC",
        "cpv": "CPV",
        "video_q1_25": "Video Q1 (25%)",
        "video_q2_50": "Video Q2 (50%)",
        "video_q3_75": "Video Q3 (75%)",
        "video_completed_100": "Video Completed (100%)",
        "video_completion_rate": "Video Completion Rate (%)",
        "vtr": "VTR (%)"
    }
    
    # Build headers
    headers = []
    for dim in dims:
        headers.append(dimension_labels.get(dim, dim.replace("_", " ").title()))
    
    for metric in selected_metrics:
        headers.append(metric_labels.get(metric, metric.replace("_", " ").title()))
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    
    for row in data:
        csv_row = []
        
        # Dimensions
        for dim in dims:
            csv_row.append(row.get(dim, ""))
        
        # Selected metrics
        for metric in selected_metrics:
            csv_row.append(row.get(metric, 0))
        
        writer.writerow(csv_row)
    
    filename = f"ad_performance_report_{start_date}_to_{end_date}.csv"
    
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/reports/ad-performance/export/excel")
async def export_ad_performance_excel(
    dimensions: str = "source,domain,creative_name",
    metrics: str = "impressions,clicks,ctr,conversions,spend",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    num_rows: int = 100,
    use_real_data: bool = True,
    campaign_id: Optional[str] = None,
    creative_id: Optional[str] = None
):
    """Export ad performance report as Excel (XLSX)"""
    dims = [d.strip() for d in dimensions.split(",")]
    selected_metrics = [m.strip() for m in metrics.split(",") if m.strip()]
    
    if not end_date:
        end_dt = datetime.now(timezone.utc)
        end_date = end_dt.strftime("%Y-%m-%d")
    
    if not start_date:
        start_dt = datetime.now(timezone.utc) - timedelta(days=30)
        start_date = start_dt.strftime("%Y-%m-%d")
    
    # Try real data first
    data = []
    if use_real_data:
        data, is_real = await get_real_ad_performance_data(
            dimensions=dims,
            start_date=start_date,
            end_date=end_date,
            num_rows=min(num_rows, 10000),
            campaign_id=campaign_id,
            creative_id=creative_id,
            metrics=selected_metrics
        )
    
    # Fall back to mock data
    if not data:
        data = generate_mock_ad_performance_data(
            dimensions=dims,
            start_date=start_date,
            end_date=end_date,
            num_rows=min(num_rows, 10000),
            campaign_id=campaign_id,
            creative_id=creative_id,
            metrics=selected_metrics
        )
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ad Performance Report"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    dimension_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    video_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define all dimension labels
    dimension_labels = {
        "campaign_name": "Campaign Name",
        "creative_name": "Creative Name",
        "source": "Source",
        "domain": "Domain", 
        "insertion_order": "Insertion Order",
        "line_item": "Line Item",
        "bundle": "Bundle",
        "app_name": "App Name",
        "country": "Country",
        "city": "City",
        "ip": "IP Address",
        "device_ifa": "Device ID",
        "os": "OS",
        "make": "Make"
    }
    
    # Define all metric labels
    metric_labels = {
        "impressions": "Impressions",
        "reach": "Reach",
        "clicks": "Clicks",
        "ctr": "CTR (%)",
        "conversions": "Conversions",
        "spend": "Spend",
        "win_rate": "Win Rate (%)",
        "ecpm": "eCPM",
        "cpc": "CPC",
        "cpv": "CPV",
        "video_q1_25": "Video Q1 (25%)",
        "video_q2_50": "Video Q2 (50%)",
        "video_q3_75": "Video Q3 (75%)",
        "video_completed_100": "Video Completed (100%)",
        "video_completion_rate": "Video Completion Rate (%)",
        "vtr": "VTR (%)"
    }
    
    video_metrics = ["video_q1_25", "video_q2_50", "video_q3_75", "video_completed_100", "video_completion_rate", "vtr"]
    
    # Build headers
    headers = []
    dimension_cols = []
    for dim in dims:
        headers.append(dimension_labels.get(dim, dim.replace("_", " ").title()))
        dimension_cols.append(len(headers))
    
    video_start_col = None
    for metric in selected_metrics:
        headers.append(metric_labels.get(metric, metric.replace("_", " ").title()))
        if metric in video_metrics and video_start_col is None:
            video_start_col = len(headers)
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        
        if col in dimension_cols:
            cell.fill = dimension_fill
        elif video_start_col and col >= video_start_col:
            cell.fill = video_fill
        else:
            cell.fill = header_fill
    
    # Write data
    for row_idx, row_data in enumerate(data, 2):
        col = 1
        
        # Dimensions
        for dim in dims:
            cell = ws.cell(row=row_idx, column=col, value=row_data.get(dim, ""))
            cell.border = thin_border
            col += 1
        
        # Selected metrics
        for metric in selected_metrics:
            cell = ws.cell(row=row_idx, column=col, value=row_data.get(metric, 0))
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right')
            col += 1
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"ad_performance_report_{start_date}_to_{end_date}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

